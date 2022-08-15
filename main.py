import json
import random
import re
from time import sleep
from typing import Generator

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

URL: str = 'https://nb-bet.com'
HEADERS: dict = {
    "Accept": '*/*',
    "Accept-Encoding": 'gzip, deflate',
    "Accept-Language": 'RU-ru, ru;q=0.5',
    "Cache-Control": 'no-cache',
    "Connection": 'keep-alive',
    "DNT": '1',
    "Host": 'nb-bet.com',
    "Pragma": 'no-cache',
    "Sec-Fetch-Site": 'cross-site',
    "TE": 'trailers',
    "Upgrade-Insecure-Requests": '1',
    "User-Agent": 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:103.0) Gecko/20100101 Firefox/103.0'
}
SESSION = requests.session()
SESSION.cookies.clear()
SESSION.get(URL, headers=HEADERS)


def get_leagues() -> Generator[tuple[str, str], None, None]:
    url: str = 'https://app.nb-bet.com/v1/page-soccer-leagues/{}/0/true/120/0/true/1/false'
    page: int = 1
    leagues: list[dict, ...] = [{}, ]
    h = HEADERS.copy()
    h["Referer"] = 'https://nb-bet.com/Tournaments'
    h = dict(
        [(k, h[k]) for k in random.choices(tuple(h.copy()), k=6)]
    )
    while leagues:
        r = SESSION.get(
            url.format(page),
            headers=dict(
                [(k, h[k]) for k in random.choices(tuple(h.copy()), k=6)]
            )
        )

        inc: int = 0
        while not r.ok:
            print(inc)
            sleep(inc * 5)
            inc += 1
            r = SESSION.get(
                url.format(page),
                headers=dict(
                    [(k, h[k]) for k in random.choices(tuple(h.copy()), k=6)]
                )
            )

        leagues = r.json()["data"]["leagues"]

        for L in leagues:
            yield L["1"], L["2"]  # Name, link

        page += 1


def get_teams(link: str) -> dict[str: str]:
    r = SESSION.get(
        f"{URL}{link}",
        headers=dict(
            [(k, HEADERS[k]) for k in random.choices(tuple(HEADERS.copy()), k=6)]
        )
    )

    if not r.ok:
        return {}

    soup = BeautifulSoup(r.text, 'html.parser')

    teams: dict[str: str] = {}

    table = soup.find('table', class_="table-summary")

    if table is None:
        return {}

    for tr in table('tr', class_=re.compile("tournaments-stats-cells")):
        team_cell = tr.find('a')
        teams[team_cell.text] = team_cell["href"][7:]  # Slicing out '/Teams/'

    return teams


def dump_table(link: str, amount: int, current_season: bool) -> Generator[int, None, None]:
    league_id: int = int(link.split('-')[0][13:])
    teams: dict = get_teams(link)
    matches: list = get_matches(link)
    wb = Workbook()

    yield 1.0
    wb.create_sheet('Игры дома', 0)
    wb['Игры дома'].append(
        [
            '№',
            'Команда', 'Кол-во игр дома',
            'Ср. ударов в створ', 'Ср. забитых голов',
            'Ср. пропущенных ударов в створ', 'Ср. пропущенных голов'
        ]
    )
    yield from dump_to_sheet(wb['Игры дома'], teams.items(),
                             lambda team: get_team_info(league_id, team, amount, current_season, True))

    yield 2.0
    wb.create_sheet('Игры в гостях', 1)
    wb['Игры в гостях'].append(
        [
            '№',
            'Команда', 'Кол-во игр дома',
            'Ср. ударов в створ', 'Ср. забитых голов',
            'Ср. пропущенных ударов в створ', 'Ср. пропущенных голов'
        ]
    )
    yield from dump_to_sheet(wb['Игры в гостях'], teams.items(),
                             lambda team: get_team_info(league_id, team, amount, current_season, False))

    yield 3.0
    wb.create_sheet('Матчи', 2)
    wb['Матчи'].append(
        [
            '№',
            'date',
            'home', 'guest',
            'Голы 1 команды', 'Голы 2 команды', 'Победитель',
            'Удары 1 команды', 'Удары 2 команды',
            'Удары в створ 1 команды', 'Удары в створ 2 команды',
            'Угловые 1 команды', 'Угловые 2 команды',
            'Атаки 1 команды', 'Атаки 2 команды',
            'Опасные атаки 1 команды', 'Опасные атаки 2 команды',
            'Жёлтые карточки 1 команды', 'Жёлтые карточки 2 команды',
            'Штрафные 1 команды', 'Штрафные 2 команды',
            'Сэйвы 1 команды', 'Сэйвы 2 команды',
            'Фолы 1 команды', 'Фолы 2 команды',
            'Вбрасывания 1 команды', 'Вбрасывания 2 команды',
            'Вероятность победы 1 команды', 'Вероятность ничьи', 'Вероятность победы 2 команды'
        ]
    )
    yield from dump_to_sheet(wb['Матчи'], matches,
                             get_match_info)

    wb.remove(wb['Sheet'])

    wb.save('Stats.xlsx')


def dump_to_sheet(sheet, args, row_func) -> Generator[int, None, None]:
    columns: list = []

    for i, arg in enumerate(args):
        columns.append(
            tuple(row_func(arg))
        )
        yield i + 1

    yield 0

    columns = sorted(columns.copy())

    for i, col in enumerate(columns):
        sheet.append(
            (i + 1,) + col
        )
        yield -(i + 1)


def get_team_info(league_id: int, team: tuple[str, str], amount: int, current_season: bool, at_home: bool) -> tuple:
    url = ("https://app.nb-bet.com/v1/page-soccer-team/"
           f"{team[1]}/{amount}/{league_id}/"
           f"{json.dumps(current_season)}/true/{json.dumps(at_home)}/{json.dumps(not at_home)}/false")

    r = SESSION.get(
        url,
        headers=dict(
            [(k, HEADERS[k]) for k in random.choices(tuple(HEADERS.copy()), k=6)]
        )
    )

    if not r.ok:
        return (team[0],) + ('',) * 5

    match_team_info = r.json()["data"]["team"]["12"][0][0]

    if not match_team_info:
        return (team[0],) + ('',) * 5

    if "20" not in match_team_info:
        return (team[0],) + ('',) * 5

    amount = match_team_info["1"]
    av_shot_on_goal = match_team_info["20"]
    av_goal = match_team_info["8"]
    av_missed_shot_on_goal = match_team_info["21"]
    av_missed_goal = match_team_info["9"]

    return team[0], amount, av_shot_on_goal, av_goal, av_missed_shot_on_goal, av_missed_goal


def get_matches(link: str) -> list:
    url: str = f'{URL}{link}'
    league_id: int = int(link.split('-')[0][13:])

    r = SESSION.get(
        url,
        headers=dict(
            [(k, HEADERS[k]) for k in random.choices(tuple(HEADERS.copy()), k=6)]
        )
    )

    soup = BeautifulSoup(r.text, 'html.parser')
    container = soup.find('div', id=f"MainContent_TabContainer_{league_id}_ctl04")
    table = container.find('table')

    return [tr.find('a')["href"] for tr in table('tr')]


def get_match_info(link: str, increment: int = 0):
    if 'live' in link:
        print('Live')
        return '',

    url = f'{URL}{link}'

    h = HEADERS.copy()
    h["Referer"] = link

    r = SESSION.get(
        url,
        headers=dict(
            [(k, h[k]) for k in random.choices(tuple(h.copy()), k=6)]
        )
    )

    if not r.ok:
        print(r.status_code, increment)
        sleep(5 * increment)
        return get_match_info(link, increment + 1)

    soup = BeautifulSoup(r.text, 'html.parser')
    container = soup.find('div', class_="MatchScoreboard__ScoreboardScoreContainer-sc-1al6574-5")

    date = container.find('div', class_="MatchScoreboard__ScoreboardDate-sc-1al6574-4").text.split(' ')[0]
    teams = tuple(i.a.text for i in container('div', class_="MatchScoreboard__ScoreboardTeamName-sc-1al6574-8"))
    scores = tuple(
        div.text
        for div in container.find('div', class_="MatchScoreboard__ScoreboardScore-sc-1al6574-11")
        ('div')
    )
    winner: int
    if scores[0] > scores[1]:
        winner = 1
    elif scores[0] < scores[1]:
        winner = 2
    else:
        winner = 0
    stats = soup.find('div', class_="MatchStats__Items-sc-157qk4v-1")
    stats_list = []
    for i in stats('div', class_="MatchStats__Item-sc-157qk4v-2")[1:-1]:
        info = i.find('div')
        if (
                info.find('div', class_="MatchStats__ItemNumberTitle-sc-157qk4v-5").text
                not in ('Удары', 'Удары в створ',
                        'Угловые',
                        'Атаки', 'Опасные атаки',
                        'Жёлтые карточки', 'Штрафные',
                        'Сэйвы', 'Фолы', 'Вбрасывания')
        ):
            continue
        stats_list.extend(
            [
                info.find('div', class_="MatchStats__ItemNumberLeft-sc-157qk4v-3").text,
                info.find('div', class_="MatchStats__ItemNumberRight-sc-157qk4v-4").text
            ]
        )
    del stats_list[1]

    probabilities = soup.find('div', class_="MatchOutcomesProbability__Legend-pl6igr-2")

    probabilities_list: list
    if probabilities is None:
        probabilities_list = []
    else:
        probabilities_list = [
            float('0.' + i.text.split()[-1][:-1])
            for i in probabilities('div', recursive=False)
        ]

    return [date, *teams, *scores, winner, *stats_list, *probabilities_list]
